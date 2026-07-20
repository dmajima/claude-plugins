#!/usr/bin/env python3
"""deep-test / test-report: Excel 報告書生成スクリプト。

test-results.yaml / test-cases.yaml を読み込み、テスト報告書（.xlsx 1 ファイル）を
openpyxl の全コード生成方式で出力する（テンプレートファイル不使用）。

- フォーマット SSOT: プラグイン共通 references/report-format.md 3 章
  （シート構成・サマリ / 推移 / レベル別の内容・列定義・スタイル・免責注記）
- 読み込み・集計・共通定数は同ディレクトリの report_model.py に一元化
  （generate_markdown.py と共有。集計規則: latest 採用 / deprecated 除外）
- 画像は埋め込まず、エビデンス参照列に相対パス文字列（改行区切り）で記載する。
  パスの基準（テスト実績データディレクトリ）はサマリシートに注記として必ず出力する
- 集計はすべて本スクリプトの機械集計であり、LLM の手計算を介在させない

使い方:
    python generate_excel.py --results <test-results.yaml> --cases <test-cases.yaml> \
        --output <出力 .xlsx パス>
"""

import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import argparse
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# 同ディレクトリの共通データモデルモジュール（report_model.py）を import する
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from report_model import (
    LEVEL_TERM_NOTES,
    VERDICT_DESCRIPTIONS,
    annotation_target,
    build_disclaimer_rows,
    build_model,
    date_part,
    evidence_path_note,
    format_duration,
    format_extras_cell,
    format_extras_value_lines,
    join_lines,
    level_label,
    load_yaml,
    manual_breakdown_note,
    sanitize,
    sanitized_count,
)

# ---------------------------------------------------------------------------
# Excel 固有の定数（report-format.md 3 章準拠）
# ---------------------------------------------------------------------------

# report-format.md 3.5 status 条件付き書式（塗りつぶし, 文字色）
STATUS_STYLES = {
    "pass": ("C6EFCE", "006100"),
    "fail": ("FFC7CE", "9C0006"),
    "blocked": ("FCE4D6", "974706"),
    "skipped": ("D9D9D9", "404040"),
    "na": ("F2F2F2", "808080"),
}

# report-format.md 3.5 ヘッダ行スタイル（濃紺背景・白字太字）
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)
THIN_SIDE = Side(style="thin")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CELL_ALIGN = Alignment(wrap_text=True, vertical="top")

# report-format.md 3.4 列定義 + 3.5 列幅目安
LEVEL_SHEET_COLUMNS = [
    ("ケース ID", 16),
    ("revision", 9),
    ("タイトル", 32),
    ("優先度", 9),
    ("実行手順", 45),
    ("期待結果", 35),
    ("実際結果", 35),
    ("status", 10),
    ("実行主体", 12),
    ("reason", 25),
    ("実行時間(秒)", 10),
    ("エビデンス参照", 40),
    ("NG 詳細", 50),
    ("補足情報（extras）", 40),
]


# ---------------------------------------------------------------------------
# 整形ヘルパ（Excel 固有）
# ---------------------------------------------------------------------------


def format_test_data(test_data):
    if isinstance(test_data, dict):
        return "\n".join(f"{k}: {v}" for k, v in test_data.items())
    return "" if test_data is None else str(test_data)


def format_ng_detail(result):
    """NG 詳細列（fail 行のみ）: 再現手順・検証データ・severity・補足情報（extras）。"""
    defect = (result or {}).get("defect") or {}
    if not defect:
        return ""
    lines = [f"severity: {defect.get('severity', '')}", "再現手順:"]
    for step in defect.get("reproduction_steps") or []:
        lines.append(str(step))
    lines.append("検証データ:")
    lines.append(format_test_data(defect.get("test_data")))
    # 補足情報（defect.extras）: レベル別の拡張情報を改行区切りで併記（長大値は切り詰め）。
    # list 値（session_findings 等）は repr のままにせず要素ごとの箇条書き行で整形する
    extras = defect.get("extras")
    if isinstance(extras, dict) and extras:
        lines.append("補足情報:")
        for key, value in extras.items():
            value_lines = format_extras_value_lines(value)
            if isinstance(value, list) and value:
                lines.append(f"{key}:")
                for item in value_lines:
                    lines.append(f"- {item}")
            else:
                lines.append(f"{key}: {value_lines[0]}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel 書き込みヘルパ
# ---------------------------------------------------------------------------


def style_header_row(ws, row_idx, num_cols):
    for offset in range(num_cols):
        cell = ws.cell(row=row_idx, column=1 + offset)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGN
        cell.border = TABLE_BORDER


def write_table(ws, start_row, headers, rows):
    """ヘッダ 1 行 + データ行を罫線・折返し付きで書き、最終行の次の行番号を返す。"""
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=sanitize(header))
    style_header_row(ws, start_row, len(headers))
    row_idx = start_row
    for row in rows:
        row_idx += 1
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=sanitize(value))
            cell.alignment = CELL_ALIGN
            cell.border = TABLE_BORDER
    return row_idx + 1


def write_section_label(ws, row_idx, text):
    ws.cell(row=row_idx, column=1, value=sanitize(text)).font = SECTION_FONT
    return row_idx + 1


def apply_status_format(ws, cell_range):
    """status 値セルへの条件付き書式（report-format.md 3.5）。"""
    for status, (bg, fg) in STATUS_STYLES.items():
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill("solid", start_color=bg, end_color=bg),
                font=Font(color=fg),
            ),
        )


def setup_page(ws, last_col, last_row, title_rows="1:1"):
    """A4 横・横 1 ページ収め・タイトル行リピート・印刷範囲（report-format.md 3.5）。

    title_rows: リピート印刷する行範囲。用語注記行を持つシート（ユニット / 単体）は
    注記行 + ヘッダ行を含めた範囲（例: "1:2"）を指定する。
    """
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = title_rows
    ws.print_area = f"A1:{get_column_letter(max(last_col, 1))}{max(last_row, 1)}"


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


# ---------------------------------------------------------------------------
# シート生成
# ---------------------------------------------------------------------------


def build_summary_sheet(ws, model, generated_on):
    ws.title = "サマリ"
    ws.cell(row=1, column=1, value=sanitize(f"テスト報告書 サマリ: {model['target']}"))
    style_header_row(ws, 1, 9)
    row = 3

    # 基本情報
    row = write_section_label(ws, row, "基本情報")
    exec_dates = sorted({date_part(r.get("executed_at")) for r in model["runs"] if r.get("executed_at")})
    if not exec_dates:
        exec_range = ""
    elif len(exec_dates) == 1:
        exec_range = exec_dates[0]
    else:
        exec_range = f"{exec_dates[0]} 〜 {exec_dates[-1]}"
    info_rows = [
        ("対象", model["target"]),
        ("報告書生成日", generated_on),
        ("実行日", exec_range),
        ("run 数", len(model["runs"])),
    ]
    row = write_table(ws, row, ["項目", "内容"], info_rows)
    # エビデンスパスの基準注記（必須出力。基準は results 実パスから導出。report_model.evidence_path_note）
    note_cell = ws.cell(
        row=row, column=1,
        value=sanitize(evidence_path_note(model["target"], model.get("results_path"))),
    )
    note_cell.alignment = CELL_ALIGN
    row += 2

    # run 情報
    row = write_section_label(ws, row, "run 情報")
    run_rows = [
        (
            r.get("run_id", ""),
            r.get("executed_at", ""),
            r.get("finished_at") or "",
            r.get("status", ""),
            r.get("mode", ""),
            len(r.get("scope") or []),
            r.get("environment", ""),
        )
        for r in model["runs"]
    ]
    row = write_table(
        ws, row,
        ["run_id", "実行日時", "終了日時", "status", "mode", "対象ケース数", "environment"],
        run_rows,
    )
    row += 1

    # レベル別集計（latest 採用。自動 / 手動 = 実行主体別内訳）
    row = write_section_label(ws, row, "レベル別集計（latest 採用）")
    agg_rows = [
        (
            level_label(lr["level"]),
            lr["target"],
            lr["pass"],
            lr["fail"],
            lr["blocked"],
            lr["skipped"],
            lr["na"],
            lr["auto"],
            lr["manual"],
        )
        for lr in model["level_rows"]
    ]
    t = model["total"]
    agg_rows.append(
        (
            "合計", t["target"], t["pass"], t["fail"], t["blocked"], t["skipped"], t["na"],
            t["auto"], t["manual"],
        )
    )
    row = write_table(
        ws, row,
        ["レベル", "対象数", "pass", "fail", "blocked", "skipped", "na", "自動", "手動"],
        agg_rows,
    )
    # 手動内訳注記（必須出力。report_model.manual_breakdown_note）
    breakdown_cell = ws.cell(row=row, column=1, value=sanitize(manual_breakdown_note(model)))
    breakdown_cell.alignment = CELL_ALIGN
    row += 2

    # NG 一覧
    row = write_section_label(ws, row, "NG 一覧（latest が fail のケース）")
    if model["ng_rows"]:
        ng_table = [
            (n["case_id"], level_label(n["level"]), n["title"], n["severity"], join_lines(n["evidence"]))
            for n in model["ng_rows"]
        ]
        row = write_table(
            ws, row, ["ケース ID", "レベル", "タイトル", "severity", "エビデンス参照"], ng_table
        )
        ws.cell(row=row, column=1, value="再現手順・検証データの詳細は各レベル別シートの NG 詳細列を参照。")
        row += 1
    else:
        ws.cell(row=row, column=1, value="なし")
        row += 1
    row += 1

    # 未確認事項（skipped 一覧 + ケース定義に存在しない実績 ID）
    row = write_section_label(ws, row, "未確認事項（skipped 一覧）")
    if model["skipped_rows"]:
        sk_table = [
            (s["case_id"], level_label(s["level"]), s["reason"]) for s in model["skipped_rows"]
        ]
        row = write_table(ws, row, ["ケース ID", "レベル", "reason"], sk_table)
    else:
        ws.cell(row=row, column=1, value="なし")
        row += 1
    if model["unknown_latest_ids"]:
        row = write_table(
            ws, row,
            ["ケース定義に存在しない実績 ID", "備考"],
            [
                (cid, "test-cases.yaml 未定義のため集計・明細の対象外（ケース定義の削除・改名等がないか確認すること）")
                for cid in model["unknown_latest_ids"]
            ],
        )
    if model["unknown_level_case_ids"]:
        row = write_table(
            ws, row,
            ["未知レベルの実施済みケース", "備考"],
            [
                (
                    cid,
                    "LEVEL_ORDER（8 値）以外の level のため unknown として集計・総合判定に反映"
                    "（ケース定義の level 値を確認すること）",
                )
                for cid in model["unknown_level_case_ids"]
            ],
        )
    if model["masked_case_ids"]:
        row = write_table(
            ws, row,
            ["機微情報マスキング適用ケース", "備考"],
            [
                (
                    cid,
                    "既知シークレットパターンを検出し evidence-policy.md 5 章の形式でマスクした"
                    "（実績・エビデンスへの生値の混入経路を確認すること）",
                )
                for cid in model["masked_case_ids"]
            ],
        )
    row += 1

    # 所見・注記（annotations。実行結果に影響しない注釈。空なら表ごと省略）
    if model["annotations"]:
        row = write_section_label(
            ws, row, "所見・注記（実行結果に影響しない注釈。annotate 追記分）"
        )
        row = write_table(
            ws, row,
            ["出所", "対象", "本文"],
            [(a["source"], annotation_target(a), a["text"]) for a in model["annotations"]],
        )
        row += 1

    # 総合判定
    row = write_section_label(ws, row, "総合判定")
    verdict_cell = ws.cell(row=row, column=1, value=model["verdict"])
    verdict_cell.font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=sanitize(VERDICT_DESCRIPTIONS[model["verdict"]])).alignment = CELL_ALIGN
    row += 1
    ws.cell(
        row=row, column=1,
        value="総合判定は実績の機械的集計であり、リリース可否・受入可否の判断は人間が行う。",
    )
    row += 2

    # 免責注記
    row = write_section_label(ws, row, "免責注記")
    row = write_table(ws, row, ["項目", "注記"], build_disclaimer_rows(model))

    set_widths(ws, [24, 22, 22, 14, 12, 14, 60, 10, 10])
    ws.freeze_panes = "A2"
    setup_page(ws, 9, row)


def build_transition_sheet(ws, model):
    ws.title = "推移"
    num_cols = max(8, 1 + len(model["runs"]))
    ws.cell(row=1, column=1, value=sanitize(f"推移: {model['target']}"))
    style_header_row(ws, 1, num_cols)
    row = 3

    # 上段: run 集計推移（時系列昇順）
    row = write_section_label(ws, row, "run 集計推移（時系列昇順）")
    run_table = [
        (
            rr["run"].get("run_id", ""),
            rr["run"].get("executed_at", ""),
            rr["run"].get("mode", ""),
            rr["counts"]["pass"],
            rr["counts"]["fail"],
            rr["counts"]["blocked"],
            rr["counts"]["skipped"],
            rr["counts"]["na"],
        )
        for rr in model["run_rows"]
    ]
    row = write_table(
        ws, row,
        ["run_id", "実行日時", "mode", "pass", "fail", "blocked", "skipped", "na"],
        run_table,
    )
    row += 1

    # 下段: ケース別推移（セル = 当該 run での status。未実行は空欄）
    row = write_section_label(ws, row, "ケース別推移（セル = 当該 run での status。未実行は空欄）")
    matrix_header = ["ケース ID"] + [r.get("run_id", "") for r in model["runs"]]
    matrix_rows = [tuple([m["case_id"]] + m["statuses"]) for m in model["matrix"]]
    header_row_idx = row
    row = write_table(ws, row, matrix_header, matrix_rows)
    if model["matrix"] and model["runs"]:
        first = header_row_idx + 1
        last = header_row_idx + len(model["matrix"])
        last_col = get_column_letter(1 + len(model["runs"]))
        apply_status_format(ws, f"B{first}:{last_col}{last}")

    set_widths(ws, [20] + [22] * max(len(model["runs"]), 7))
    ws.freeze_panes = "A2"
    setup_page(ws, num_cols, row)


def build_level_sheet(wb, model, level):
    ws = wb.create_sheet(level_label(level))
    headers = [c[0] for c in LEVEL_SHEET_COLUMNS]
    # ユニット / 単体シートのみ: 誤読防止の用語注記をシート先頭（表の直上）に 1 行出力する
    # （report-format.md 3.4。他レベルのシートには不要）
    term_note = LEVEL_TERM_NOTES.get(level)
    start_row = 1
    if term_note:
        note_cell = ws.cell(row=1, column=1, value=sanitize(term_note))
        note_cell.alignment = CELL_ALIGN
        note_cell.font = SECTION_FONT
        start_row = 2
    rows = []
    for case in model["active_cases"]:
        if case.get("level") != level:
            continue
        cid = case.get("id")
        detail = model["latest_detail"].get(cid)
        if not detail:
            continue  # 実行結果明細のため未実行ケースは対象外（対象数はサマリで把握）
        ref = detail["ref"]
        result = detail["result"]
        duration = result.get("duration_sec")
        rows.append(
            (
                cid,
                ref.get("case_revision", ""),
                case.get("title", ""),
                case.get("priority", ""),
                join_lines(case.get("steps")),
                case.get("expected", ""),
                result.get("actual") or "",
                ref.get("status", ""),
                result.get("executed_by") or "",
                result.get("reason") or "",
                format_duration(duration),
                join_lines(result.get("evidence")),
                format_ng_detail(result),
                format_extras_cell(result.get("extras")),
            )
        )
    last_row = write_table(ws, start_row, headers, rows) - 1
    if rows:
        apply_status_format(ws, f"H{start_row + 1}:H{last_row}")
    set_widths(ws, [c[1] for c in LEVEL_SHEET_COLUMNS])
    ws.freeze_panes = f"B{start_row + 1}"  # （注記行 +）ヘッダ行 + ケース ID 列固定
    setup_page(ws, len(headers), last_row, title_rows=f"1:{start_row}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="deep-test テスト報告書（Excel）生成")
    parser.add_argument("--results", required=True, help="test-results.yaml のパス")
    parser.add_argument("--cases", required=True, help="test-cases.yaml のパス")
    parser.add_argument("--output", required=True, help="出力 .xlsx のパス")
    args = parser.parse_args()

    results_doc = load_yaml(args.results, "test-results.yaml")
    cases_doc = load_yaml(args.cases, "test-cases.yaml")
    model = build_model(cases_doc, results_doc)
    # エビデンスパス基準注記の導出元（results 実パス。非既定 base でも実配置と一致させる）
    model["results_path"] = args.results

    generated_on = date.today().strftime("%Y-%m-%d")
    wb = Workbook()
    build_summary_sheet(wb.active, model, generated_on)
    build_transition_sheet(wb.create_sheet(), model)
    for level in model["executed_levels"]:
        build_level_sheet(wb, model, level)

    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(args.output)

    if sanitized_count():
        print(f"[WARN] 禁止記号（U+00A7）を {sanitized_count()} 箇所で代替表現に置換しました")
    total = model["total"]
    sheets = ["サマリ", "推移"] + [level_label(lv) for lv in model["executed_levels"]]
    print(f"[DONE] Excel 報告書を生成しました: {args.output}")
    print(f"シート: {', '.join(sheets)}")
    print(f"総合判定: {model['verdict']}")
    print(
        "集計（latest）: 対象 {0} / pass {1} / fail {2} / blocked {3} / skipped {4} / na {5}".format(
            total["target"], total["pass"], total["fail"],
            total["blocked"], total["skipped"], total["na"],
        )
    )
    print(
        f"NG 件数: {len(model['ng_rows'])} / 未確認事項（skipped）: {len(model['skipped_rows'])}"
        f" / 所見・注記: {len(model['annotations'])}"
    )


if __name__ == "__main__":
    main()
